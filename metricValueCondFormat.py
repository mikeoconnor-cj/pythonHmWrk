#! python3
# metricValueCondFormat.py   # second version of the script

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# wb = load_workbook('/Users/michaeloconnor/Downloads/metricValOperDashMbrMoDevPublish20200205.xlsx')
wb = load_workbook('/Users/michael.oconnor/mbrMoPvt_63.xlsx')
# /Users/michael.oconnor/Downloads/QAdateAnalysis_Prod_20201016.xlsx'
# worksheet_names = wb.sheetnames # this variable is presented as a list, so we can manipulate it accordingly, but finding the index number using the name.
# print(worksheet_names)
# sheet_index = worksheet_names.index('mbr mo v3') # this will return the index number where your worksheet resides.  It should be mentioned that this sheet name must exist in the worksheet_names list variable
# wb.active = sheet_index

# ws = wb.active
ws = wb['Sheet']
# df = pd.read_excel('/Users/michaeloconnor/Downloads/metricValOperDashMbrMoDevPublish20200205.xlsx')
df = pd.read_excel('/Users/michael.oconnor/mbrMoPvt_63.xlsx', sheet_name='Sheet')
    # the above throws an error in python3 about a missinig dependancy on xlrs.  
# /Users/michael.oconnor/Downloads/QAdateAnalysis_Prod_20201016.xlsx
# # print(df)
#
# df1 = df[['org_id', 'org_level_category_cd', 'org_group_id', 'Jul-19', 'Aug-19', 'Sep-19', 'Oct-19', 'Nov-19', 'Dec-19', 'Jan-20', 'Feb-20', 'Mar-20', 'Apr-20', 'May-20', 'Jun-20', 'Jul-20', 'Aug-20', 'Sep-20']]
# , 'm-2021-08', 'm-2021-09'
df1 = df[['ORG_ID', 'm-2019-01','m-2019-02','m-2019-03','m-2019-04','m-2019-05','m-2019-06','m-2019-07','m-2019-08','m-2019-09','m-2019-10','m-2019-11','m-2019-12','m-2020-01','m-2020-02','m-2020-03','m-2020-04','m-2020-05','m-2020-06','m-2020-07','m-2020-08','m-2020-09','m-2020-10','m-2020-11','m-2020-12','m-2021-01','m-2021-02','m-2021-03','m-2021-04','m-2021-05','m-2021-06', 'm-2021-07', 'm-2021-08', 'm-2021-09']]
# , 'm-2021-01','m-2021-02','m-2021-03'

# df1 = df[['org_id', 'org_level_category_cd', 'org_group_id', 'q-2018-1', 'q-2018-2', 'q-2018-3', 'q-2018-4', 'q-2019-1', 'q-2019-2'
#     , 'q-2019-3', 'q-2019-4', 'q-2020-1']]

print(df1)

# https://thispointer.com/select-rows-columns-by-name-or-index-in-dataframe-using-loc-iloc-python-pandas/

# set all cells to red, save to a new test file
rows = dataframe_to_rows(df1, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, col in enumerate(row, 1):
        #ws.cell(row=r_idx, column=c_idx, value=col).font = Font(color=colors.RED)
        #ws[r_idx,c_idx] = 'test'  #column in excel are letters

        if r_idx > 1 and c_idx > 2:
            curVal = df1.iloc[r_idx - 2, c_idx - 1]  #need to adjust indexes to 0 based, excluding header rw
            priorVal = df1.iloc[r_idx - 2, c_idx - 2]
            absDiff = abs(curVal - priorVal) * 1.0  # convert to float

            # checkCalc = absDiff / priorVal # runtime warning: invalid value encountered in double_scalars
            #avoid divide by zero and get the 0 prior val to ### curVal excluding the 0 prior and 0 cur case
            if (priorVal > 0 and absDiff / priorVal > .05) or (absDiff == curVal and absDiff != 0):
            # if df1.iloc[r_idx - 2, c_idx - 1] == 44415 and df1.iloc[r_idx - 2, c_idx - 2] == 44415:
                # cell = ws.cell(row=r_idx, column=c_idx, value=col)
                cell = ws.cell(row=r_idx, column=c_idx)
                if curVal < priorVal:
                    # cell.font = Font(color=colors.RED)
                    # cell.background = Background(color=colors.RED)
                    # cell.fill = PatternFill(fgColor=colors.RED, fill_type="solid") # works in python 2

                    cell.fill = PatternFill(fgColor='00FF0000', fill_type="solid")
                    # cell.fill = PatternFill(patternType="solid",bgColor=colors.RED)
                    # cell.fill = PatternFill()
                elif curVal > priorVal:
                    # cell.font = Font(color=colors.GREEN)
                    # cell.background = Background(color=colors.GREEN)
                    # cell.fill = PatternFill(patternType="solid",fgColor=colors.BLACK,bgColor=colors.GREEN)

                    # cell.fill = PatternFill(fgColor=colors.GREEN, fill_type="solid") # works in python 2

                    cell.fill = PatternFill(fgColor='0000FF00', fill_type="solid")
        # Data can be assigned directly to cells
#         #ws['A1'] = 42
#
# # wb.save('/Users/michael.oconnor/Documents/memberMoAttribFormattedProd_202003.xlsx')
wb.save('/Users/michael.oconnor/mbrMoPvt_45_formatted.xlsx')
# wb.save('/Users/michael.oconnor/pmpy_2_formatted.xlsx')
# /Users/michael.oconnor/Downloads/QAdateAnalysis_Prod_20201016_v02.xlsx