# MDPCPupdatePractIDsInRosters.py

# a script to revert the practice IDs to the ones 
# the practices had upon entry into the MCPCP program
# and that match the data in the claims file types

from pathlib2 import Path
from openpyxl import load_workbook
import re
import time

# testing breakpoints.. debug not running python script ..and not breaking
# for i in range(3):
#     print(i)

practiceIDsToRevertInRosters = ['T2MD0261'
                            ,'T2MD0274'
                            ,'T2MD0287'
                            ,'T2MD0288'
                            ,'T2MD0307'
                            ,'T2MD0313'
                            ,'T2MD0316'
                            ,'T2MD0365'
                            ,'T2MD0368'
                            ,'T2MD0369'
                            ,'T2MD0370'
                            ,'T2MD0372'
                            ,'T2MD0415'
                            ,'T2MD0421'
                            ,'T2MD0427'
                            ,'T2MD0428'
                            ,'T2MD0459'
                            ,'T2MD0466'
                            ,'T2MD0467'
                            ,'T2MD0470'
                            ,'T2MD0472'
                            ,'T2MD0478'
                            ,'T2MD0481'
                            ,'T2MD0482'
                            ,'T2MD0514'
                            ,'T2MD0515'
                            ,'T2MD0518'
                            ,'T2MD0522'
                            ,'T2MD0524'
                            ,'T2MD0555'
                            ,'T2MD0564'
                            ,'T2MD0568'
                            ,'T2MD0574'
                            ,'T2MD0575'
                            ,'T2MD0625'
                            ,'T2MD0641'
                            ,'T2MD0695'
                            ,'T2MD0698'
                            ,'T2MD0703'
                            ,'T2MD0710'
                            ,'T2MD0723'
                            ,'T2MD0724'
                            ,'T2MD0727'
                            ,'T2MD0744'
]

tDaydate = time.strftime("%Y%m%d")


p = Path("D:\\Users\\michael.oconnor\\q3_2022\\attrib")

myObjList = list(p.glob('*_BeneAttrRpt_*'))
# print(myObjList)
# [WindowsPath('D:/Users/michael.oconnor/q3_2022/attrib/T1MD0939_BeneAttrRpt_CY2022_Q3_20220729.xlsx'), WindowsPath('D:/Users/michael.oconnor/q3_2022/attrib/T2MD0316_BeneAttrRpt_CY2022_Q3_20220729.xlsx')]

workSheets = ['Attributed Beneficiaries', 'Bene Eligibility Recoupment']
parser = re.compile(r'(T\dMD\d\d\d\d)(.*)')


for i in range(len(myObjList)):    
    wb = load_workbook(str(myObjList[i]))

    mo = parser.search(str(myObjList[i])) 
    if mo == None:
        print('no match')
        continue
    else:
        print(mo.group(0), mo.group(1))  # shows 1) full file name and 
                                                 # 2) the practice ID from the filename

    if mo.group(1) not in practiceIDsToRevertInRosters:
        continue

    
    # check for T1 here... else continue
    for j in range(len(workSheets)):
        ws = wb[workSheets[j]]
        practID = ws['B6'].value
        practIDRevert = practID.replace('T2','T1')
        print('Practice Id in worksheet ' + workSheets[j] + ': ' + practID + ' ModifiedPID: ' + practIDRevert)
        ws['B6'].value = practIDRevert


    # print('A date for the new files: ' + tDaydate)

    # could have Regexed this out
    hardCodedPieceToReplace = '20220729'
    fName = mo.group(0).replace(hardCodedPieceToReplace, tDaydate)

    fnameAndPathToSav = p.joinpath(fName)
    # print(type(fnameToSav))
    # <class 'pathlib2.WindowsPath'>
    print(fnameAndPathToSav)  # can't save this object: pathlib2.WindowsPath
    # D:\Users\michael.oconnor\q3_2022\attrib\T2MD0316_BeneAttrRpt_CY2022_Q3_20220818.xlsx

    # maybe save here instead? D:\Users\michael.oconnor\q3_2022\attribModifd

    wb.save(str(fnameAndPathToSav))
    wb.close()

