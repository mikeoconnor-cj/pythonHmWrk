# read in the mbr mo grp by xlsx, open another wb, 
import openpyxl

# THE QUERY
# USE WAREHOUSE INT_A3327;
# USE DATABASE INT_A3327;
# USE WAREHOUSE PROD_MEDSTAR_MDPCP;
# USE DATABASE prod_medstar_mdpcp; 

# select
#     org_id
# --  , org_level_category_cd
# --  , org_group_id
#   , period_id
#   , measure_value_decimal
# from insights.metric_value_operational_dashboard
# --SNAPSHOT_PROD_MEDSTAR_MDPCP_20201118.insights.metric_value_operational_dashboard
# where measure_cd = 'total_member_years_current_month'
# 	and patient_medicare_group_cd = '#NA' 
# 	and org_level_category_cd = 'aco'
# 	and attribution_type = 'as_was'
# 	and substr(period_id,3,7) >= '2019-01' 
# order by org_id    
#     	, substr(period_id,3,7)

#  SELECT 'SELECT org_id, period_id, measure_value_decimal ' 
 
# || 'FROM ' || database_name || '.' || 'insights.metric_value_operational_dashboard'
# || ' WHERE measure_cd = ''' || 'total_member_years_current_month' 
# || ''' and patient_medicare_group_cd = ''' || '#NA'
# || ''' and org_level_category_cd = ''' || 'aco'
# || ''' and attribution_type = ''' || 'as_was'
# || ''' and substr(period_id,3,7) >= ''' || '2019-01'
# --|| ''' order_by org_id, substr(period_id,3,7)'
# || ''' UNION ALL '  AS queryPart
# -- 
# -- || '''' 
# --|| SPLIT_PART(database_name,'_',2) || ''' AS ORG_ID, ' 
# --|| ''''  || tbl.table_schema || '.' || tbl.table_name || ''' AS TableName, ' 
# --|| 'MAX(Load_Ts) AS MaxLoadTS, Count(*) as rwCount FROM ' || database_name || '.' || tbl.table_schema || '.' || tbl.table_name	|| ';'

# FROM information_schema.DATABASES
# JOIN information_schema.TABLES tbl 
# 	ON 1 = 1
# JOIN information_schema.COLUMNS col
# 	ON tbl.TABLE_NAME = col.TABLE_NAME 
# 	AND tbl.TABLE_SCHEMA  = col.TABLE_SCHEMA 
# 	AND col.COLUMN_NAME = 'LOAD_TS'
# WHERE DATABASE_NAME LIKE 'PROD_%' AND SPLIT_PART(database_name,'_',3) = ''
# --DATABASE_NAME NOT LIKE '%_FE'
# 	AND tbl.TABLE_SCHEMA = 'INSIGHTS'
# 						--'VRDC'	
# 	AND tbl.table_name = 'METRIC_VALUE_OPERATIONAL_DASHBOARD'
# 	AND SPLIT_PART(database_name,'_',2) IN (
# 			'A1052'
# 			, 'A2024'
# 			, 'A2251'
# 			, 'A2575'
# 			, 'A2841'
# 			, 'A3229'
# 			, 'A3320'
# 			, 'A3327'
# 			, 'A3367'
# 			, 'A3599'
# 			, 'A3632'
# 			, 'A3667'
# 			, 'A3669'
# 			, 'A3774'
# 			, 'A3822'
# 			, 'A4585'
# 			, 'A4588'
# 			, 'A4709'
# 			, 'A4768'
# 			, 'A4806'
# 			, 'A4824'
# 	) 
# --ORDER BY DATABASE_NAME 	
# union all
# select 'order by org_id, substr(period_id,3,7)' as queryPart  
# 
# another measure: ' WHERE measure_cd = ''' || 'total_pmpy_current_month' 
#
# 
# MDPCP with member months at 'at_time_tin' org_level
# Select ORG_GROUP_ID  	
# --  , ORG_GROUP_NAME 
#  , period_id	
#   , measure_value_decimal AS totlForGroup	
# from insights.metric_value_operational_dashboard	
# where measure_cd = 'total_member_years_current_month'	
# 	and patient_medicare_group_cd = '#NA' 
# 	and org_level_category_cd = 'at_time_tin'
# 	and attribution_type = 'as_was'
# 	and substr(period_id,3,7) >= '2019-01' 
# order by ORG_GROUP_ID 
# --	, ORG_GROUP_NAME  	
# 	,  period_id       



wb = openpyxl.load_workbook('/Users/michael.oconnor/Downloads/mbrMosGrpbyFormat_3.xlsx')
# wb = openpyxl.load_workbook('/Users/michael.oconnor/Downloads/mbrMosGrpbyFormat_mdpcp.xlsx')
sheet = wb['Sheet37']
# sheet = wb['snapshot']

wb2 = openpyxl.Workbook()
sheet2 = wb2.active

orgData = {}
yearMoData = {}
orgYrMoStruct = {}


# 'MD': {'Allegany': {'pop': 75087, 'tracts': 23},
#         'Anne Arundel': {'pop': 537656, 'tracts': 104},
#         'Baltimore': {'pop': 805029, 'tracts': 214},
#         'Baltimore City': {'pop': 620961, 'tracts': 200},



# read the rows and populate a dict, to get a distict list of orgs
for rowNum in range(2, sheet.max_row + 1):
    # alternative
    # org = sheet.cell(row=rowNum, column=1).value
    org = sheet['A' + str(rowNum)].value
    yrMo = sheet['B' + str(rowNum)].value
    pmpy = sheet['C' + str(rowNum)].value

    orgData.setdefault(org, {})  # how else can you create 
                                # distinct lists?
    yearMoData.setdefault(yrMo, {})

    orgYrMoStruct.setdefault(org, {})
    orgYrMoStruct[org].setdefault(yrMo, {'pmpyNo': 0})

    orgYrMoStruct[org][yrMo]['pmpyNo'] += pmpy

orgDictKeyList = orgData.keys()
orgList = list(orgDictKeyList)

yrMoDictKeyList = yearMoData.keys()
yrMoList = sorted(list(yrMoDictKeyList))
# sorted year Mo list - how?
# for idx, val in enumerate(yrMoList):  # idx starts w/ 0
#     print(idx)
#     print(val)

# print('Done..')

# iterate through orgYrMoStruct, double for-loop?, need the key
# can I get an imperfect Excel cross tab?
# for orgItem in orgYrMoStruct:  # just returns an org string
#     for orgs in orgItem.keys():  # so there are no keys()
#         print(org)

# for orgKey in orgYrMoStruct:
#     o = orgKey
#     oVal = orgYrMoStruct[orgKey]

 # https://realpython.com/iterate-through-dictionary-python/
 # try a membership test using yrMoList's values against a
 # org's yearMonth
 # does yrMoList support enumerate so I can get the index too? using it for xlsx column  
    #  yes, starts w/ 0 
# need python 3 for this? 3.7 installed and set as interpreter.  now program can't find openpyxl
# sorted_income = {k: incomes[k] for k in sorted(incomes)}
sorted_orgYrMoStruct = {k: orgYrMoStruct[k] for k in sorted(orgYrMoStruct)}

# header 
    # create a number of dims var for the cell column parameter?
    # if I'm going to have org | orgName:
    # listOfStrings = 'A2024 | DVACO'.split('|')
    # listOfStrings = [item.strip() for item in listOfStrings]   
rwInx = 1
rwLabel1 = sheet.cell(row=rwInx, column=1).value  #org_id
sheet2.cell(row=rwInx, column=1).value = rwLabel1

for colInx, month in enumerate(yrMoList):
    sheet2.cell(row=rwInx, column=colInx + 2).value = month

rwInx += 1
for key, value in sorted_orgYrMoStruct.items():
# for key, value in orgYrMoStruct.items():
    print(key)
    print(value)
    print(str(rwInx))
    sheet2.cell(row = rwInx, column = 1).value = key
    # sorted year Mo list - how?  see above sorted
    for colInx, month in enumerate(yrMoList):
        if month in orgYrMoStruct[key].keys():
            print(str(orgYrMoStruct[key][month]['pmpyNo']))
            print(str(colInx + 2))   # could add an else and print 0
            sheet2.cell(row = rwInx, column = colInx + 2).value = orgYrMoStruct[key][month]['pmpyNo']
    rwInx += 1

# save the new wb
wb2.save('/Users/michael.oconnor/mbrMoPvt_43.xlsx')