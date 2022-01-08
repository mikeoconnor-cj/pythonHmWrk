# recreation 

from pathlib import Path
import re, os
import bs4, requests
import openpyxl


# using beautiful soup 

# 1st landing page: https://www.politico.com/tag/cartoon-carousel
# landingPageFile = open('https://www.politico.com/tag/cartoon-carousel')
res = requests.get('https://www.politico.com/tag/cartoon-carousel')
res.raise_for_status()

landingPageFileSoup = bs4.BeautifulSoup(res.text, 'html.parser')

# latest page in carosel - first one found.  
elems = landingPageFileSoup.select('img[data-lazy-img]')
# elems = landingPageFileSoup.select('a[img="data-lazy-img"]')
# for elem in elems: 
#     print(elem)
#     print(elem.parent)
#     print(elem.parent.get('href'))
print(elems[0].parent.get('href'))
newLatestHref = elems[0].parent.get('href')
parseSequenceNo = re.compile(r'\d\d\d\d\d\d')
mo = parseSequenceNo.search(elems[0].parent.get('href'))
# print(mo.group(2))
newLatestSeqStr = mo.group(0)


# update the log excel file
curWD = Path.cwd()

filePth = curWD / 'politicoURLlog.xlsx'
wb = openpyxl.load_workbook(filePth)
ws = wb.active
curHref = ws.cell(row=2, column=1).value
curSeq = ws.cell(row=2, column=2).value

# write latest data to spreadsheet
if curHref != newLatestHref: 
    ws.cell(row=2, column=1).value = newLatestHref
    ws.cell(row=2, column=2).value = newLatestSeqStr
    strFilePth = str(filePth)
    wb.save(strFilePth)

wb.close()


print(elems[0].parent.get('href'))
res2 = requests.get(elems[0].parent.get('href'))
res2.raise_for_status()

latestComicsSoup = bs4.BeautifulSoup(res2.text, 'html.parser')

latestComicElems = latestComicsSoup.select('img[data-lazy-img]')

parseLatestComicFname = re.compile(r'F\d\d*-.*\.(jpg|jpeg|pdf|gif|png|tiff|bmp|svg)')

for elem in latestComicElems:
    # print(elem.getText())
    # print(elem.attrs)
    # print(elem['data-lazy-img'])
    print('Downloading image %s...' % (elem['data-lazy-img']))
    mtchObj = parseLatestComicFname.search(elem['data-lazy-img'])
    print(mtchObj.group(0))
    imgURLfileName = mtchObj.group(0)
    pathFname = '/politico/' + imgURLfileName
    imgURLfilePth = open(os.path.abspath(curWD) + pathFname, 'wb')

    res3 = requests.get(elem['data-lazy-img'])
    res3.raise_for_status()
    for chunk in res3.iter_content(100000):
        imgURLfilePth.write(chunk)
    imgURLfilePth.close() 

# # print(mo.group(2))
# flickrImgUrl = mo.group(2)
# print('Downloading image %s...' % (flickrImgUrl))
# res = requests.get(flickrImgUrl)
# res.raise_for_status()

# # Save the image to ./flickr
# imageFile = open(os.path.join('flickr', os.path.basename(flickrImgUrl)),'wb')
# for chunk in res.iter_content(100000):
#     imageFile.write(chunk)
# imageFile.close()


# ~/Downloads/pythonHmWrk/politico  # this doesn't work
# file iterator on target directory, ? use glob, use unlink()
p = Path('/Users/michaeloconnor/Downloads/pythonHmWrk/politico')
myPPObjList = list(p.glob('*'))  # PosixPaths

for indx, FQfName in enumerate(myPPObjList):
    print(str(indx), FQfName)

# delete the old files


# save the new ones .. parse out filenames using regex
# F\d\d*-.*\.(jpg|jpeg|pdf|gif|png|tiff|bmp|svg)


# remember to save to git hub to protect the script.  
