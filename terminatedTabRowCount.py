from pathlib2 import Path
from openpyxl import load_workbook
import pandas as pd

# D:\Users\michael.oconnor\q2\attribRptsUnzipd

p = Path('D:/Users/michael.oconnor/q2/attribRptsUnzipd')
# q2/allClaimsUnzipped
myObjList = list(p.glob('*'))

numRows = 0

for i in range(len(myObjList)):  # is there another way to get an index and the item ?enumerate?
    # print(str(i) + ' ' + str(myObjList[i]))
    wb = load_workbook(str(myObjList[i]))
    ws = wb["Bene Eligibility Recoupment"]
    practcName = ws['B6'].value
    numRows = ws.max_row
    colA = ws['A']
    lenColA = len(colA)

    df = pd.read_excel(str(myObjList[i]), sheet_name="Bene Eligibility Recoupment", skiprows=8)
    # print(df)
    df1 = df[["Last Name", "First Name"]]
    df1 = df1.dropna()
    # print(practcName + '\t' + str(numRows) + '\t' + str(lenColA))
    # numRows = 0
    # print(practcName)
    # print(df1)
    print(practcName + '\t' + str(len(df1["Last Name"])))
    # print("")
    wb.close()
