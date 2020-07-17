from pathlib2 import Path
from openpyxl import load_workbook

p = Path('D:/Users/michael.oconnor/q1/q1_2020addtnlClaimsUnzp')
# q2/allClaimsUnzipped
myObjList = list(p.glob('*_stat_*'))
# print(myObjList[0])
# print(str(myObjList[0]))

# wb = load_workbook(str(myObjList[0]))
# ws = wb.active
# fileName = ws['A6'].value
# numRecs = ws['B6'].value
# print(fileName + '\t' + str(numRecs))
# wb.close()

for i in range(len(myObjList)):  # is there another way to get an index and the item ?enumerate?
    # print(str(i) + ' ' + str(myObjList[i]))
    wb = load_workbook(str(myObjList[i]))
    ws = wb.active
    fileName = ws['A6'].value
    numRecs = ws['C6'].value
    print(fileName + '\t' + str(numRecs))
    wb.close()
