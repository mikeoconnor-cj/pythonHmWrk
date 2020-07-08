import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# wb = load_workbook('/Users/michaeloconnor/Downloads/metricValOperDashMbrMoDevPublish20200205.xlsx')
wb = load_workbook('/Users/michael.oconnor/Documents/QAdateAnalysis_Prod_20200625_vdi.xlsx')
# worksheet_names = wb.sheetnames # this variable is presented as a list, so we can manipulate it accordingly, but finding the index number using the name.
# print(worksheet_names)
# sheet_index = worksheet_names.index('mbr mo v3') # this will return the index number where your worksheet resides.  It should be mentioned that this sheet name must exist in the worksheet_names list variable
# wb.active = sheet_index

ws = wb.active
# df = pd.read_excel('/Users/michaeloconnor/Downloads/metricValOperDashMbrMoDevPublish20200205.xlsx')
df = pd.read_excel('/Users/michael.oconnor/Documents/QAdateAnalysis_Prod_20200625_vdi.xlsx')
# # print(df)
#
df1 = df[['org_id', 'org_level_category_cd', 'org_group_id', 'Jan-19', 'Feb-19', 'Mar-19', 'Apr-19', 'May-19', 'Jun-19'
    , 'Jul-19', 'Aug-19', 'Sep-19', 'Oct-19', 'Nov-19', 'Dec-19', 'Jan-20', 'Feb-20', 'Mar-20', 'Apr-20', 'May-20', 'Jun-20']]

# df1 = df[['org_id', 'org_level_category_cd', 'org_group_id', 'q-2018-1', 'q-2018-2', 'q-2018-3', 'q-2018-4', 'q-2019-1', 'q-2019-2'
#     , 'q-2019-3', 'q-2019-4', 'q-2020-1']]

print(df1)

# https://thispointer.com/select-rows-columns-by-name-or-index-in-dataframe-using-loc-iloc-python-pandas/

# set all cells to red, save to a new test file
rows = dataframe_to_rows(df1, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, col in enumerate(row, 1):
        #ws.cell(row=r_idx, column=c_idx, value=col).font = Font(color=colors.RED)
        #ws[r_idx,c_idx] = 'test'  #column in excel are letters

        if r_idx > 1 and c_idx > 4:
            curVal = df1.iloc[r_idx - 2, c_idx - 1]  #need to adjust indexes to 0 based, excluding header rw
            priorVal = df1.iloc[r_idx - 2, c_idx - 2]
            absDiff = abs(curVal - priorVal)
            #avoid divide by zero and get the 0 prior val to ### curVal excluding the 0 prior and 0 cur case
            if (priorVal > 0 and absDiff / priorVal > .05) or (absDiff == curVal and absDiff != 0):
            # if df1.iloc[r_idx - 2, c_idx - 1] == 44415 and df1.iloc[r_idx - 2, c_idx - 2] == 44415:
                # cell = ws.cell(row=r_idx, column=c_idx, value=col)
                cell = ws.cell(row=r_idx, column=c_idx)
                if curVal < priorVal:
                    # cell.font = Font(color=colors.RED)
                    # cell.background = Background(color=colors.RED)
                    cell.fill = PatternFill(fgColor=colors.RED, fill_type="solid")
                    # cell.fill = PatternFill(patternType="solid",bgColor=colors.RED)
                    # cell.fill = PatternFill()
                elif curVal > priorVal:
                    # cell.font = Font(color=colors.GREEN)
                    # cell.background = Background(color=colors.GREEN)
                    # cell.fill = PatternFill(patternType="solid",fgColor=colors.BLACK,bgColor=colors.GREEN)
                    cell.fill = PatternFill(fgColor=colors.GREEN, fill_type="solid")
        # Data can be assigned directly to cells
#         #ws['A1'] = 42
#
# # wb.save('/Users/michael.oconnor/Documents/memberMoAttribFormattedProd_202003.xlsx')
wb.save('/Users/michael.oconnor/Documents/QAdateAnalysis_Prod_20200625_vdi.xlsx')